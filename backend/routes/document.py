import os
import logging
from uuid import uuid4
from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from werkzeug.utils import secure_filename
from backend.config import Config
from backend.database.db import SessionLocal, mongo_db
from backend.models.models import Document, AnalysisHistory
from backend.services.parser import DocumentParser
from backend.services.vector_store import VectorStoreManager
from backend.agents.agent import FinancialIntelligenceAgent
from backend.exceptions import AnalysisStorageError, DocumentParseError
logger = logging.getLogger(__name__)

doc_bp = Blueprint('document', __name__)

ALLOWED_EXTENSIONS = {'pdf', 'csv', 'xlsx', 'xls'}


def _remove_file(file_path):
    if not os.path.exists(file_path):
        return
    try:
        os.remove(file_path)
    except OSError:
        logger.warning("Unable to remove file %s.", file_path, exc_info=True)


def _mark_document_failed(document_id):
    failed_db = SessionLocal()
    try:
        failed_doc = failed_db.query(Document).filter(Document.id == document_id).first()
        if failed_doc:
            failed_doc.status = "failed"
            failed_db.commit()
    except Exception:
        failed_db.rollback()
        logger.exception("Unable to mark document %s as failed.", document_id)
    finally:
        failed_db.close()

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def check_is_financial_file(file_path, file_type):
    """Classifies uploaded document content to ensure it is a valid financial document."""
    text = ""
    try:
        if file_type == "PDF":
            import fitz
            doc = fitz.open(file_path)
            pages_to_read = min(len(doc), 5)
            extracted = []
            for i in range(pages_to_read):
                t = doc[i].get_text()
                if t:
                    extracted.append(t)
            doc.close()
            text = "\n".join(extracted)
        elif file_type == "CSV":
            with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                lines = [f.readline() for _ in range(40)]
                text = "\n".join(lines)
        elif file_type == "Excel":
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True)
            sheets = wb.sheetnames
            if sheets:
                ws = wb[sheets[0]]
                rows = []
                for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
                    if r_idx > 25:
                        break
                    rows.append(" ".join([str(val) for val in row if val is not None]))
                text = "\n".join(rows)
            wb.close()
    except Exception as e:
        logger.exception("Unable to inspect uploaded file %s during validation.", file_path)
        raise DocumentParseError(
            f"Unable to read uploaded {file_type.lower()} file: {e}",
            user_message=(
                f"The uploaded {file_type.lower()} file could not be read; it may be corrupt or unreadable."
            )
        ) from e

    if not text or not text.strip():
        return False

    text_lower = text.lower()

    # Explicit Non-Financial / Resume / Academic / Certificate indicators
    non_financial_keywords = [
        "curriculum vitae", "resume", "work experience", "education", "hobbies",
        "objective", "academic background", "personal details", "declaration",
        "skills & abilities", "technical skills", "projects undertaken", "certificate of",
        "certifies that", "degree of", "university", "bachelor of", "master of", "diploma in",
        "schooling", "gpa", "cgpa", "semester", "personal profile", "employment history",
        "job summary", "passed the examination", "coursework", "references available",
        "cover letter", "dear sir/madam", "application for", "hall ticket", "admit card"
    ]

    non_fin_matches = [kw for kw in non_financial_keywords if kw in text_lower]

    # Explicit Financial Document keywords & structure indicators
    financial_keywords = [
        "revenue", "expense", "profit", "loss", "income", "balance sheet", "assets", "liabilities", "equity",
        "cash flow", "financial statement", "transaction", "gst", "invoice", "payroll", "ledger", "account",
        "cogs", "sales", "ebitda", "capital", "retained earning", "tax", "audit", "credit", "debit", "deposit", 
        "withdrawal", "price", "amount", "cost", "margin", "liquidity", "debt", "quick ratio", "current ratio",
        "bank statement", "account statement", "opening balance", "closing balance", "subtotal", "bill to",
        "invoice date", "tax invoice", "statement period", "net profit", "gross profit", "operating cash flow"
    ]

    fin_matches = [kw for kw in financial_keywords if kw in text_lower]

    # If document has resume/academic/certificate indicators and weak financial context, reject
    if len(non_fin_matches) >= 2 or (len(non_fin_matches) >= 1 and len(fin_matches) < 3):
        logger.info("Rejected non-financial document with validation indicators: %s", non_fin_matches)
        return False

    # Require at least 2 distinct financial keywords
    if len(fin_matches) < 2:
        logger.info("Rejected document with insufficient financial keywords: %s", fin_matches)
        return False

    return True

@doc_bp.route('/upload', methods=['POST'])
@doc_bp.route('/document/upload', methods=['POST'])
@jwt_required()
def upload_file():
    user_id = get_jwt_identity()

    if 'file' not in request.files:
        return jsonify({"message": "No file part in the request"}), 400
        
    file = request.files['file']
    if file.filename == '':
        return jsonify({"message": "No selected file"}), 400

    if not allowed_file(file.filename):
        return jsonify({
            "message": "This platform only accepts financial documents such as invoices, bank statements, balance sheets, profit and loss statements, cash flow reports, and other business financial records."
        }), 400

    # Ensure upload folder exists
    try:
        os.makedirs(Config.UPLOAD_FOLDER, exist_ok=True)
    except OSError:
        logger.exception("Unable to prepare upload folder %s.", Config.UPLOAD_FOLDER)
        return jsonify({"message": "Unable to prepare storage for the uploaded file. Please try again."}), 500

    filename = secure_filename(file.filename)
    if not filename or not allowed_file(filename) or '.' not in filename:
        return jsonify({"message": "The uploaded filename is invalid."}), 400

    try:
        numeric_user_id = int(user_id)
    except (TypeError, ValueError):
        return jsonify({"message": "Invalid authenticated user"}), 401

    user_upload_folder = os.path.join(Config.UPLOAD_FOLDER, str(numeric_user_id))
    try:
        os.makedirs(user_upload_folder, exist_ok=True)
    except OSError:
        logger.exception("Unable to prepare upload folder %s.", user_upload_folder)
        return jsonify({"message": "Unable to prepare storage for the uploaded file. Please try again."}), 500

    file_path = os.path.join(user_upload_folder, f"{uuid4().hex}_{filename}")
    try:
        file.save(file_path)
        file_size = os.path.getsize(file_path)
    except Exception:
        logger.exception("Unable to save or inspect uploaded file %s.", filename)
        if os.path.exists(file_path):
            try:
                os.remove(file_path)
            except OSError:
                logger.warning("Unable to remove incomplete upload %s.", file_path, exc_info=True)
        return jsonify({"message": "Unable to save the uploaded file. Please try again."}), 500

    # Validate file size
    if file_size > 10 * 1024 * 1024:  # 10 MB limit
        try:
            os.remove(file_path)
        except OSError:
            logger.warning("Unable to remove oversized upload %s.", file_path, exc_info=True)
        return jsonify({"message": "File exceeds the 10MB size limit."}), 413

    # Determine file type category
    ext = filename.rsplit('.', 1)[1].lower()
    file_type = "PDF"
    if ext == 'csv':
        file_type = "CSV"
    elif ext in ['xlsx', 'xls']:
        file_type = "Excel"

    # content-based validation check before parser/db record/embeddings/analysis
    try:
        is_financial_file = check_is_financial_file(file_path, file_type)
    except DocumentParseError as e:
        logger.warning("Uploaded file failed content validation: %s", e, exc_info=True)
        try:
            os.remove(file_path)
        except OSError:
            logger.warning("Unable to remove unreadable upload %s.", file_path, exc_info=True)
        return jsonify({"message": e.user_message}), 422

    if not is_financial_file:
        if os.path.exists(file_path):
            try:
                os.remove(file_path)
            except OSError:
                logger.warning("Unable to remove rejected upload %s.", file_path, exc_info=True)
        return jsonify({
            "message": "This platform only accepts financial documents such as invoices, bank statements, balance sheets, profit and loss statements, cash flow reports, and other business financial records."
        }), 400

    # Initialize MySQL DB session and create record since file is validated
    db = SessionLocal()
    doc_record = None
    try:
        doc_record = Document(
            user_id=numeric_user_id,
            filename=filename,
            file_path=file_path,
            status="processing",
            file_type=file_type
        )
        db.add(doc_record)
        db.commit()
        db.refresh(doc_record)

        # 1. Document Parsing & Text Extraction
        chunks = []
        if file_type == "PDF":
            chunks = DocumentParser.parse_pdf(file_path)
        elif file_type == "CSV":
            chunks = DocumentParser.parse_csv(file_path)
        elif file_type == "Excel":
            chunks = DocumentParser.parse_excel(file_path)

        if not chunks:
            doc_record.status = "failed"
            db.commit()
            _remove_file(file_path)
            return jsonify({"message": "Document contains no extractable data or is corrupted."}), 400

        # 2. Chunking & Embeddings
        vstore = VectorStoreManager.get_instance()
        vstore.clear()  # Clear workspace cache
        vstore.add_chunks(chunks)

        # 3. LangChain Agent Execution
        agent = FinancialIntelligenceAgent()
        analysis_result = agent.run_analysis(doc_record.id)

        return jsonify({
            "success": True,
            "current_state_analysis": analysis_result.get("current_state_analysis", {}),
            "gap_detection": analysis_result.get("gap_detection", []),
            "forward_looking_flags": analysis_result.get("forward_looking_flags", []),
            "metrics": analysis_result.get("metrics", {}),
            "charts": analysis_result.get("charts", {}),
            "analysis_mode": analysis_result.get("analysis_mode", "llm"),
            "degraded_reason": analysis_result.get("degraded_reason"),
            "metrics_source": analysis_result.get("metrics_source"),
            "data_quality": analysis_result.get("data_quality"),
            "pdf_url": f"/api/report/{doc_record.id}",
            "document": {
                "id": doc_record.id,
                "filename": doc_record.filename,
                "file_type": doc_record.file_type,
                "status": "analyzed"
            }
        }), 200

    except DocumentParseError as e:
        db.rollback()
        logger.warning("Document %s could not be parsed: %s", filename, e, exc_info=True)
        if doc_record:
            _mark_document_failed(doc_record.id)
        _remove_file(file_path)
        return jsonify({"message": e.user_message}), 422
    except AnalysisStorageError:
        db.rollback()
        logger.exception("Analysis for document %s could not be saved.", filename)
        if doc_record:
            _mark_document_failed(doc_record.id)
        return jsonify({"message": "Analysis could not be saved. Please try again."}), 500
    except Exception:
        db.rollback()
        logger.exception("Unexpected error during upload/analysis processing for %s.", filename)
        if doc_record:
            _mark_document_failed(doc_record.id)
        _remove_file(file_path)
        return jsonify({"message": "An unexpected server error occurred. Please try again."}), 500
    finally:
        db.close()

@doc_bp.route('/history', methods=['GET'])
@jwt_required()
def get_history():
    user_id = get_jwt_identity()
    db = SessionLocal()
    try:
        docs = db.query(Document).filter(Document.user_id == int(user_id)).order_by(Document.created_at.desc()).all()
        history_list = []
        for doc in docs:
            history_list.append({
                "id": doc.id,
                "filename": doc.filename,
                "file_type": doc.file_type,
                "status": doc.status,
                "created_at": doc.created_at.isoformat()
            })
        return jsonify({"history": history_list}), 200
    except Exception:
        logger.exception("Unexpected error while loading document history.")
        return jsonify({"message": "An unexpected server error occurred. Please try again."}), 500
    finally:
        db.close()

@doc_bp.route('/history/<int:doc_id>', methods=['GET'])
@jwt_required()
def get_history_detail(doc_id):
    user_id = get_jwt_identity()
    db = SessionLocal()
    try:
        doc = db.query(Document).filter(Document.id == doc_id, Document.user_id == int(user_id)).first()
        if not doc:
            return jsonify({"message": "Document not found"}), 404
            
        result = mongo_db["analysis_results"].find_one({"document_id": doc_id})
        if not result:
            return jsonify({"message": "Analysis details not found"}), 404
            
        if '_id' in result:
            result['_id'] = str(result['_id'])
            
        return jsonify({"analysis": result, "filename": doc.filename}), 200
    except Exception:
        logger.exception("Unexpected error while loading analysis details for document %s.", doc_id)
        return jsonify({"message": "An unexpected server error occurred. Please try again."}), 500
    finally:
        db.close()

@doc_bp.route('/report/<int:doc_id>', methods=['GET'])
@jwt_required()
def download_report(doc_id):
    user_id = get_jwt_identity()
    db = SessionLocal()
    try:
        doc = db.query(Document).filter(Document.id == doc_id, Document.user_id == int(user_id)).first()
        if not doc:
            return jsonify({"message": "Document not found"}), 404
            
        pdf_name = f"report_{doc_id}.pdf"
        pdf_path = os.path.join(Config.REPORTS_FOLDER, pdf_name)
        
        if not os.path.exists(pdf_path):
            return jsonify({"message": "PDF report file not found on server."}), 404
            
        return send_file(
            pdf_path,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f"Financial_Analysis_{doc.filename}.pdf"
        )
    except Exception:
        logger.exception("Unexpected error while downloading report for document %s.", doc_id)
        return jsonify({"message": "An unexpected server error occurred. Please try again."}), 500
    finally:
        db.close()

@doc_bp.route('/chat', methods=['POST'])
@doc_bp.route('/document/chat', methods=['POST'])
@jwt_required()
def chat_with_document():
    data = request.get_json()
    if not data or 'message' not in data or 'document_id' not in data:
        return jsonify({"message": "Missing message or document_id"}), 400
        
    doc_id = data['document_id']
    message = data['message']
    history = data.get('history', [])
    user_id = get_jwt_identity()
    
    db = SessionLocal()
    try:
        doc = db.query(Document).filter(Document.id == doc_id, Document.user_id == int(user_id)).first()
        if not doc:
            return jsonify({"message": "Document not found or unauthorized"}), 404
            
        agent = FinancialIntelligenceAgent()
        response_text = agent.run_chat_query(doc_id, message, history)
        
        return jsonify({
            "success": True,
            "response": response_text
        }), 200
    except Exception:
        logger.exception("Unexpected error while answering document chat request.")
        return jsonify({"message": "An unexpected server error occurred. Please try again."}), 500
    finally:
        db.close()

@doc_bp.route('/document/delete/<int:doc_id>', methods=['DELETE'])
@jwt_required()
def delete_document(doc_id):
    user_id = get_jwt_identity()
    db = SessionLocal()
    try:
        doc = db.query(Document).filter(Document.id == doc_id, Document.user_id == int(user_id)).first()
        if not doc:
            return jsonify({"message": "Document not found or unauthorized"}), 404
            
        filename = doc.filename
        file_path = doc.file_path
        pdf_name = f"report_{doc_id}.pdf"
        pdf_path = os.path.join(Config.REPORTS_FOLDER, pdf_name)
        
        # 1. Delete database records in MySQL
        db.query(AnalysisHistory).filter(AnalysisHistory.document_id == doc_id).delete()
        db.delete(doc)
        db.commit()
        
        cleanup_failures = []

        # 2. Delete analysis from MongoDB
        try:
            mongo_db["analysis_results"].delete_one({"document_id": doc_id})
        except Exception:
            cleanup_failures.append("analysis record")
            logger.exception("Failed to delete MongoDB analysis for document %s.", doc_id)

        # 3. Delete physical files if they exist
        if os.path.exists(file_path):
            try:
                os.remove(file_path)
            except OSError:
                cleanup_failures.append("uploaded file")
                logger.warning("Failed to delete uploaded file %s.", file_path, exc_info=True)
                
        if os.path.exists(pdf_path):
            try:
                os.remove(pdf_path)
            except OSError:
                cleanup_failures.append("PDF report")
                logger.warning("Failed to delete PDF report %s.", pdf_path, exc_info=True)
                
        # 4. Remove document chunks from vector store
        try:
            vstore = VectorStoreManager.get_instance()
            vstore.remove_document_chunks(filename)
        except Exception:
            cleanup_failures.append("vector store chunks")
            logger.exception("Failed to remove vector-store chunks for document %s.", doc_id)
        
        return jsonify({
            "success": True,
            "message": (
                "Document deleted successfully"
                if not cleanup_failures
                else f"Document deleted from the database, but cleanup failed for: {', '.join(cleanup_failures)}."
            )
        }), 200
        
    except Exception:
        db.rollback()
        logger.exception("Unexpected error deleting document %s.", doc_id)
        return jsonify({"message": "An unexpected server error occurred. Please try again."}), 500
    finally:
        db.close()
